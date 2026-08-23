from __future__ import annotations

import argparse
import json
import os
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from build_sidecar import REPOSITORY_ROOT, rust_target_triple
from openpyxl import load_workbook  # type: ignore[import-untyped]


def run_frozen_sidecar(
    sidecar_path: Path,
    input_text: str,
    *,
    environment: dict[str, str],
    timeout: int,
) -> subprocess.CompletedProcess[str]:
    """Run the UTF-8 NDJSON sidecar contract independently of host locale."""
    return subprocess.run(
        [str(sidecar_path)],
        input=input_text,
        capture_output=True,
        check=False,
        env=environment,
        text=True,
        encoding="utf-8",
        errors="strict",
        timeout=timeout,
    )


def decode_protocol_output(raw: bytes) -> str:
    """Decode protocol bytes strictly as UTF-8 for locale-independent tests."""
    return raw.decode("utf-8", errors="strict")


def request(request_id: str, method: str, params: dict[str, Any]) -> str:
    return json.dumps(
        {
            "protocol_version": "0.1",
            "request_id": request_id,
            "method": method,
            "params": params,
        },
        separators=(",", ":"),
    )


def verify(sidecar_path: Path) -> None:
    if not sidecar_path.is_file():
        raise FileNotFoundError(f"Frozen sidecar does not exist: {sidecar_path}")

    environment = os.environ.copy()
    environment.pop("PYTHONHOME", None)
    environment.pop("PYTHONPATH", None)
    environment["PATH"] = ""
    with tempfile.TemporaryDirectory(prefix="scope-frozen-verify-") as temporary:
        project_parent = Path(temporary)
        source_path = project_parent / "fixture.txt"
        source_path.write_text(
            "基层治理需要政策支持。基层治理需要实践检验。", encoding="utf-8"
        )
        input_lines = [
            request("frozen-describe", "system.describe", {}),
            request("frozen-profiles", "stopwords.profiles", {}),
            request(
                "frozen-create",
                "project.create",
                {"name": "冻结验证项目", "parent_path": str(project_parent)},
            ),
        ]
        completed = run_frozen_sidecar(
            sidecar_path,
            "\n".join(input_lines) + "\n",
            environment=environment,
            timeout=30,
        )
        if completed.returncode != 0:
            raise AssertionError(
                f"Frozen sidecar exited with {completed.returncode}: {completed.stderr}"
            )
        messages = [json.loads(line) for line in completed.stdout.splitlines() if line]
        by_id = {
            message["request_id"]: message
            for message in messages
            if message["type"] == "result"
        }
        if "frozen-profiles" not in by_id:
            raise AssertionError(
                f"stopwords.profiles failed in frozen sidecar: {messages}"
            )
        profiles = by_id["frozen-profiles"]["result"]["profiles"]
        expected_profiles = {
            "scope-cn-general-v1": ("SCOPE 中文通用停用词表 v1", 86, "draft"),
            "goto456-general": ("goto456 中文通用停用词表", 746, "reference"),
            "hit": ("哈工大停用词表", 749, "reference"),
            "baidu": ("百度停用词表", 1395, "reference"),
            "scu": ("四川大学停用词表", 860, "reference"),
            "none": ("不使用停用词", 0, "reference"),
            "project-custom": ("项目自定义", 0, "reference"),
        }
        indexed_profiles = {profile["profile_id"]: profile for profile in profiles}
        if set(indexed_profiles) != set(expected_profiles):
            raise AssertionError(
                f"Unexpected frozen profiles: {sorted(indexed_profiles)}"
            )
        for profile_id, (label, count, status) in expected_profiles.items():
            profile = indexed_profiles[profile_id]
            if (profile["label"], profile["count"], profile["status"]) != (
                label,
                count,
                status,
            ):
                raise AssertionError(f"Unexpected profile {profile_id}: {profile}")
            if not profile["hash"]:
                raise AssertionError(f"Profile {profile_id} has no hash")
        capabilities = by_id["frozen-describe"]["result"]["capabilities"]
        required_capabilities = {
            "system.describe",
            "project.create",
            "project.open",
            "corpus.import_txt",
            "document.get",
            "text.clean.preview",
            "text.clean.execute",
            "text.clean.batch",
            "text.tokenize.preview",
            "text.tokenize.execute",
            "text.tokenize.batch",
            "tokenization.dictionary.import",
            "stopwords.profiles",
            "stopwords.get",
            "stopwords.resolve",
            "stopwords.import",
            "frequency.analyze",
            "frequency.latest",
            "frequency.export",
            "diagnostic.run",
            "diagnostic.crash",
            "request.cancel",
        }
        if not required_capabilities.issubset(capabilities):
            raise AssertionError(
                f"Missing frozen capabilities: {sorted(required_capabilities - set(capabilities))}"
            )

        project_path = by_id["frozen-create"]["result"]["project"]["project_path"]
        resolve_input = request(
            "frozen-resolve",
            "stopwords.resolve",
            {
                "project_path": project_path,
                "base_profile_id": "scope-cn-general-v1",
                "custom_additions": ["验证词"],
                "custom_exclusions": ["的"],
            },
        )
        import_input = request(
            "frozen-import",
            "corpus.import_txt",
            {"project_path": project_path, "file_paths": [str(source_path)]},
        )
        follow_up = run_frozen_sidecar(
            sidecar_path,
            resolve_input + "\n" + import_input + "\n",
            environment=environment,
            timeout=30,
        )
        if follow_up.returncode != 0:
            raise AssertionError(
                f"Frozen stopword resolve/import failed: {follow_up.stderr}"
            )
        follow_messages = {
            message["request_id"]: message
            for message in (
                json.loads(line) for line in follow_up.stdout.splitlines() if line
            )
        }
        resolved = follow_messages["frozen-resolve"]["result"]["profile"]
        if (
            "验证词" not in resolved["resolved_stopwords"]
            or "的" in resolved["resolved_stopwords"]
        ):
            raise AssertionError(
                f"Frozen stopword resolve produced wrong set: {resolved}"
            )
        document_id = follow_messages["frozen-import"]["result"]["entries"][0][
            "document"
        ]["document_id"]
        workflow_input = (
            "\n".join(
                [
                    request(
                        "frozen-clean",
                        "text.clean.execute",
                        {
                            "project_path": project_path,
                            "document_id": document_id,
                            "rules": {},
                        },
                    ),
                ]
            )
            + "\n"
        )
        workflow = run_frozen_sidecar(
            sidecar_path,
            workflow_input,
            environment=environment,
            timeout=30,
        )
        if workflow.returncode != 0:
            raise AssertionError(f"Frozen cleaning failed: {workflow.stderr}")
        # Each request is handled by a fresh process in this verifier; project state is persisted on disk.
        tokenization = run_frozen_sidecar(
            sidecar_path,
            request(
                "frozen-tokenize",
                "text.tokenize.execute",
                {
                    "project_path": project_path,
                    "document_id": document_id,
                    "config": {},
                },
            )
            + "\n",
            environment=environment,
            timeout=30,
        )
        if tokenization.returncode != 0:
            raise AssertionError(f"Frozen tokenization failed: {tokenization.stderr}")
        frequency = run_frozen_sidecar(
            sidecar_path,
            request(
                "frozen-frequency", "frequency.analyze", {"project_path": project_path}
            )
            + "\n",
            environment=environment,
            timeout=30,
        )
        if frequency.returncode != 0:
            raise AssertionError(f"Frozen frequency failed: {frequency.stderr}")
        frequency_message = next(
            json.loads(line) for line in frequency.stdout.splitlines() if line
        )
        frequency_result = frequency_message["result"]
        if (
            not frequency_result["rows"]
            or frequency_result["manifest"]["effective_token_count"] <= 0
        ):
            raise AssertionError(
                f"Frozen frequency result is empty: {frequency_result}"
            )
        for format_name in ("csv", "xlsx"):
            destination = project_parent / f"冻结结果.{format_name}"
            export = run_frozen_sidecar(
                sidecar_path,
                request(
                    f"frozen-export-{format_name}",
                    "frequency.export",
                    {
                        "project_path": project_path,
                        "destination": str(destination),
                        "format": format_name,
                    },
                )
                + "\n",
                environment=environment,
                timeout=30,
            )
            if export.returncode != 0 or not destination.is_file():
                raise AssertionError(
                    f"Frozen {format_name} export failed: {export.stdout} {export.stderr}"
                )
            if format_name == "xlsx":
                workbook = load_workbook(destination, read_only=True, data_only=True)
                try:
                    if workbook.sheetnames != ["词频结果", "分析说明"]:
                        raise AssertionError(
                            f"Frozen XLSX has wrong sheets: {workbook.sheetnames}"
                        )
                    expected_header = (
                        "词语",
                        "词频（TF）",
                        "文档频率（DF）",
                        "文档覆盖率",
                        "标准化词频（每万词，RF10K）",
                    )
                    actual_header = next(workbook["词频结果"].values)
                    if actual_header != expected_header:
                        raise AssertionError(
                            f"Frozen XLSX has wrong header: {actual_header}"
                        )
                finally:
                    workbook.close()

    # Keep diagnostic crash verification separate because it intentionally terminates the process.
    input_lines = [
        request("frozen-describe", "system.describe", {}),
        request("frozen-diagnostic", "diagnostic.run", {"steps": 2, "delay_ms": 0}),
    ]
    completed = run_frozen_sidecar(
        sidecar_path,
        "\n".join(input_lines) + "\n",
        environment=environment,
        timeout=20,
    )
    if completed.returncode != 0:
        raise AssertionError(
            f"Frozen sidecar exited with {completed.returncode}: {completed.stderr}"
        )

    messages = [json.loads(line) for line in completed.stdout.splitlines() if line]
    describe = [
        message for message in messages if message["request_id"] == "frozen-describe"
    ]
    diagnostic = [
        message for message in messages if message["request_id"] == "frozen-diagnostic"
    ]
    if len(describe) != 1 or describe[0]["result"]["engine_version"] != "0.0.0":
        raise AssertionError(f"Unexpected system.describe output: {describe}")

    progress = [
        message["progress"]["current"]
        for message in diagnostic
        if message["type"] == "progress"
    ]
    terminal = [message for message in diagnostic if message["type"] == "result"]
    if progress != [1, 2]:
        raise AssertionError(f"Unexpected progress sequence: {progress}")
    if len(terminal) != 1:
        raise AssertionError(f"Expected one diagnostic result: {terminal}")
    manifest = terminal[0]["result"]["reproducibility_manifest"]
    if manifest["parameters"] != {"steps": 2, "delay_ms": 0}:
        raise AssertionError(f"Unexpected frozen parameters: {manifest['parameters']}")
    if manifest["network_used"] is not False:
        raise AssertionError("Frozen diagnostic unexpectedly reported network use")

    crashed = run_frozen_sidecar(
        sidecar_path,
        request("frozen-crash", "diagnostic.crash", {}) + "\n",
        environment=environment,
        timeout=20,
    )
    if crashed.returncode == 0:
        raise AssertionError("Frozen diagnostic crash unexpectedly exited successfully")


def default_sidecar_path() -> Path:
    executable_suffix = ".exe" if os.name == "nt" else ""
    return (
        REPOSITORY_ROOT
        / "apps/desktop/src-tauri/binaries"
        / f"scope-engine-dev-{rust_target_triple()}{executable_suffix}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Verify the frozen SCOPE sidecar contract"
    )
    parser.add_argument("sidecar", nargs="?", type=Path)
    arguments = parser.parse_args()
    sidecar = arguments.sidecar or default_sidecar_path()
    verify(sidecar.resolve())
    print(f"Frozen sidecar verified: {sidecar}")


if __name__ == "__main__":
    main()
