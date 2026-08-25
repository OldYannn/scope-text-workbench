from __future__ import annotations

import csv
import json
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from scope_engine.stopword_validation import ValidationError, run_validation


class StopwordValidationTest(unittest.TestCase):
    def _prepare_config(self, root: Path) -> tuple[Path, Path]:
        fixtures = Path(__file__).parent / "fixtures" / "validation"
        corpus_root = root / "private-corpus"
        shutil.copytree(fixtures, corpus_root)
        output_path = root / "validation-output"
        config_path = root / "config.local.json"
        config_path.write_text(
            json.dumps(
                {
                    "corpora": {
                        "policy": {"path": str(corpus_root / "policy"), "notes": str(corpus_root)},
                        "interview": {"path": str(corpus_root / "interview")},
                        "academic": {"path": str(corpus_root / "academic")},
                    },
                    "output_path": str(output_path),
                    "top_n": 3,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        return config_path, corpus_root

    def test_generates_a_private_deterministic_three_corpus_review_harness(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config_path, corpus_root = self._prepare_config(root)
            before = {
                path.relative_to(corpus_root).as_posix(): path.read_bytes()
                for path in corpus_root.rglob("*.txt")
            }

            first = run_validation(config_path)
            output = Path(first["output_path"])
            self.assertEqual(
                sorted(path.name for path in output.iterdir()),
                sorted(
                    [
                        "academic_comparison.csv",
                        "corpus_summary.csv",
                        "interview_comparison.csv",
                        "manual_review_matrix.csv",
                        "policy_comparison.csv",
                        "stopword-validation.xlsx",
                        "validation-manifest.json",
                        "validation-summary.md",
                        "watchlist.csv",
                    ]
                ),
            )
            self.assertEqual(
                before,
                {
                    path.relative_to(corpus_root).as_posix(): path.read_bytes()
                    for path in corpus_root.rglob("*.txt")
                },
            )
            for csv_path in output.glob("*.csv"):
                self.assertTrue(csv_path.read_bytes().startswith(b"\xef\xbb\xbf"), csv_path.name)

            with (output / "policy_comparison.csv").open(
                encoding="utf-8-sig", newline=""
            ) as stream:
                policy_rows = list(csv.DictReader(stream))
            self.assertEqual(
                {row["停用词配置"] for row in policy_rows},
                {"No Stopwords", "SCOPE Draft", "goto456"},
            )
            self.assertEqual(
                [row["词语"] for row in policy_rows if row["停用词配置"] == "No Stopwords"],
                sorted(
                    [row["词语"] for row in policy_rows if row["停用词配置"] == "No Stopwords"],
                    key=lambda value: (
                        next(
                            int(item["词频（TF）"])
                            for item in policy_rows
                            if item["停用词配置"] == "No Stopwords" and item["词语"] == value
                        )
                        * -1
                    ),
                ),
            )
            self.assertNotIn(
                "的",
                [row["词语"] for row in policy_rows if row["停用词配置"] == "SCOPE Draft"],
            )

            with (output / "manual_review_matrix.csv").open(
                encoding="utf-8-sig", newline=""
            ) as stream:
                review_rows = list(csv.DictReader(stream))
            self.assertEqual(
                {
                    "Potential False Positive",
                    "Potential False Negative",
                    "SCOPE vs goto456 Difference",
                },
                {row["候选类型"] for row in review_rows},
            )
            self.assertTrue(all(not row["人工判断"] and not row["人工备注"] for row in review_rows))
            self.assertIn("goto456 only", {row["停用词差异"] for row in review_rows})

            with (output / "watchlist.csv").open(encoding="utf-8-sig", newline="") as stream:
                watchlist_rows = list(csv.DictReader(stream))
            self.assertEqual(30, len(watchlist_rows))
            absent = next(
                row for row in watchlist_rows if row["语料"] == "policy" and row["词语"] == "表示"
            )
            self.assertEqual("absent", absent["状态"])
            self.assertEqual("0", absent["词频（TF）"])

            workbook = load_workbook(
                output / "stopword-validation.xlsx", read_only=True, data_only=True
            )
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "语料概览",
                        "政策_NoStop",
                        "政策_SCOPE",
                        "政策_goto456",
                        "访谈_NoStop",
                        "访谈_SCOPE",
                        "访谈_goto456",
                        "学术_NoStop",
                        "学术_SCOPE",
                        "学术_goto456",
                        "人工复核矩阵",
                        "争议词观察",
                        "运行清单",
                    ],
                )
                self.assertEqual(
                    next(workbook["政策_SCOPE"].values),
                    (
                        "排名",
                        "词语",
                        "词频（TF）",
                        "文档频率（DF）",
                        "文档覆盖率",
                        "标准化词频（每万词，RF10K）",
                    ),
                )
                self.assertEqual(
                    list(workbook["政策_SCOPE"].values)[1],
                    (1, "协商", 1, 1, 1, 1111.111111111111),
                )
            finally:
                workbook.close()

            manifest = json.loads((output / "validation-manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(False, manifest["network"])
            self.assertEqual(86, manifest["stopword_configs"]["scope_draft"]["resolved_count"])
            self.assertEqual("draft", manifest["stopword_configs"]["scope_draft"]["status"])
            self.assertEqual(
                "bf8b03b9d3709222804ae89578156d1a0d8bf2b2",
                manifest["stopword_configs"]["goto456"]["pinned_source_commit"],
            )
            self.assertNotIn(str(root), json.dumps(manifest, ensure_ascii=False))

            for path in output.iterdir():
                if path.suffix in {".csv", ".md", ".json"}:
                    content = path.read_text(
                        encoding="utf-8-sig" if path.suffix == ".csv" else "utf-8"
                    )
                    self.assertNotIn(str(root), content)
                    self.assertNotIn("根据政策安排，基层治理需要通过协商进行实施", content)

            second = run_validation(config_path)
            self.assertNotEqual(first["run_id"], second["run_id"])
            self.assertEqual(first["analytical_hashes"], second["analytical_hashes"])

    def test_invalid_utf8_fails_explicitly_without_partial_output(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            config_path, corpus_root = self._prepare_config(root)
            (corpus_root / "academic" / "invalid.txt").write_bytes(b"\xff\xfe")
            with self.assertRaisesRegex(ValidationError, "UTF-8"):
                run_validation(config_path)
            self.assertFalse((root / "validation-output").exists())


if __name__ == "__main__":
    unittest.main()
