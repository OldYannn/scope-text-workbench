from __future__ import annotations

import hashlib
import json
import runpy
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from scope_engine.frequency import (
    analyze_documents,
    export_csv,
    export_xlsx,
    optimization_candidates,
)
from scope_engine.project_store import (
    clean_execute,
    create_project,
    frequency_execute,
    import_txt,
    tokenize_execute,
)
from scope_engine.stopwords import (
    available_profiles,
    duplicate_lines,
    is_eligible_token,
    resolve_stopwords,
)


class FrequencyAnalysisTest(unittest.TestCase):
    def fixture(self):
        return [
            {
                "document_id": "doc-a",
                "tokens": [{"token": token} for token in ["的", "党", "目的地", "1", "，", "党"]],
            },
            {
                "document_id": "doc-b",
                "tokens": [{"token": token} for token in ["的", "党", "政策"]],
            },
            {"document_id": "doc-c", "tokens": None},
        ]

    def test_exact_matching_and_eligibility(self):
        result = analyze_documents(self.fixture())
        self.assertEqual([row["token"] for row in result["rows"]], ["党", "1", "政策", "目的地"])
        self.assertEqual(result["manifest"]["raw_token_count"], 9)
        self.assertEqual(result["manifest"]["eligible_token_count"], 8)
        self.assertEqual(result["manifest"]["effective_token_count"], 6)
        self.assertEqual(next(row for row in result["rows"] if row["token"] == "党")["df"], 2)
        self.assertTrue(is_eligible_token("单"))
        self.assertTrue(is_eligible_token("1"))
        self.assertFalse(is_eligible_token("，"))

    def test_profile_layering_and_stable_hash(self):
        first = resolve_stopwords(additions=["项目词"], exclusions=["的"])
        second = resolve_stopwords(additions=["项目词"], exclusions=["的"])
        self.assertIn("项目词", first["resolved_stopwords"])
        self.assertNotIn("的", first["resolved_stopwords"])
        self.assertEqual(first["resolved_stopword_hash"], second["resolved_stopword_hash"])
        self.assertIn("党", "党")

    def test_candidates_and_exports(self):
        result = analyze_documents(self.fixture())
        candidates = optimization_candidates(result, coverage_threshold=0.5)
        self.assertEqual(candidates[0]["token"], "党")
        with tempfile.TemporaryDirectory() as directory:
            csv_path = export_csv(result, Path(directory) / "词频.csv")
            self.assertTrue(csv_path.read_bytes().startswith(b"\xef\xbb\xbf"))
            with csv_path.open(encoding="utf-8-sig") as stream:
                self.assertEqual(
                    stream.readline().strip(),
                    "词语,词频（TF）,文档频率（DF）,文档覆盖率,标准化词频（每万词，RF10K）",
                )
            xlsx_path = export_xlsx(result, Path(directory) / "词频.xlsx")
            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["词频结果", "分析说明"])
                rows = list(workbook["词频结果"].values)
                self.assertEqual(
                    rows[0],
                    (
                        "词语",
                        "词频（TF）",
                        "文档频率（DF）",
                        "文档覆盖率",
                        "标准化词频（每万词，RF10K）",
                    ),
                )
                self.assertEqual(rows[1][0:3], ("党", 3, 2))
                info = dict(workbook["分析说明"].values)
                self.assertEqual(
                    info["RF10K definition"],
                    "RF10K(w) = TF(w) / EffectiveTokenCount * 10000",
                )
                self.assertIn("实际参与本次统计", info["EffectiveTokenCount"])
            finally:
                workbook.close()

    def test_project_frequency_accepts_profile_config(self):
        fixture_path = Path(__file__).parent / "fixtures" / "corpus" / "frequency-gui.txt"
        with tempfile.TemporaryDirectory() as directory:
            project_path = create_project("频率配置", directory)["project"]["project_path"]
            imported = import_txt(project_path, [str(fixture_path)])
            document = imported["entries"][0]["document"]
            rules = {
                "normalize_whitespace": True,
                "normalize_newlines": True,
                "remove_urls": True,
                "strip_html": True,
                "punctuation_mode": "keep",
            }
            clean_execute(project_path, document["document_id"], rules)
            tokenize_execute(project_path, document["document_id"], {})
            result = frequency_execute(
                project_path,
                {
                    "base_profile_id": "scope-cn-general-v1",
                    "custom_additions": [],
                    "custom_exclusions": [],
                },
            )
            row = next(row for row in result["rows"] if row["token"] == "需要")
            self.assertEqual(row["tf"], 2)
            self.assertEqual(row["df"], 1)

    def test_upstream_snapshots_and_provenance_are_deterministic(self):
        root = Path(__file__).parents[1] / "src" / "scope_engine" / "resources" / "stopwords"
        expected_counts = {
            "goto456-general.txt": 746,
            "hit.txt": 749,
            "baidu.txt": 1395,
            "scu.txt": 860,
        }
        metadata = json.loads((root / "profiles.json").read_text(encoding="utf-8"))
        expected_raw_hashes = {
            "goto456-general.txt": "5c8d5dd24906615de61ae4056f9261b6fb9f42f58bc75f442fe1032b511dc04b",
            "hit.txt": "84e526454db0245cab0d167df067f00298d271ad2c86391d45f5e880c422cbae",
            "baidu.txt": "b11ff810ee5c8934dc46b57f3a1ba85457e3893e89acdafb5cd286570fe793a3",
            "scu.txt": "2c325256276f2c4ed5ec076178c08494af7a46bf44d0b9be2fc0214d5b606d41",
        }
        expected_raw_duplicates = {
            "goto456-general.txt": 0,
            "hit.txt": 18,
            "baidu.txt": 0,
            "scu.txt": 116,
        }
        for filename, count in expected_counts.items():
            words = (root / filename).read_text(encoding="utf-8").splitlines()
            self.assertEqual(len(words), count)
            self.assertEqual(words, sorted(set(words)))
            self.assertFalse(duplicate_lines(filename))
            raw = (root / "upstream" / filename).read_bytes()
            self.assertEqual(hashlib.sha256(raw).hexdigest(), expected_raw_hashes[filename])
            key = filename.removesuffix(".txt")
            self.assertEqual(metadata[key]["raw_sha256"], expected_raw_hashes[filename])
            self.assertEqual(metadata[key]["unique_token_count"], count)
            raw_words = [line.strip() for line in raw.decode("utf-8-sig").splitlines()]
            self.assertEqual(len(raw_words), metadata[key]["raw_line_count"])
            self.assertEqual(
                len(raw_words) - len(set(raw_words)), expected_raw_duplicates[filename]
            )
        with (root / "provenance.tsv").open(encoding="utf-8") as provenance:
            self.assertEqual(sum(1 for _ in provenance) - 1, 2312)
        before = (root / "provenance.tsv").read_bytes()
        runpy.run_path(str(root / "generate_provenance.py"), run_name="__main__")
        self.assertEqual(before, (root / "provenance.tsv").read_bytes())
        scope = next(
            item for item in available_profiles() if item["profile_id"] == "scope-cn-general-v1"
        )
        self.assertEqual(scope["status"], "draft")
        self.assertEqual(scope["count"], 86)
        self.assertIn("已经", resolve_stopwords()["resolved_stopwords"])
        self.assertIn("已經", resolve_stopwords()["resolved_stopwords"])


if __name__ == "__main__":
    unittest.main()
